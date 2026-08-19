"""
HubSpot Duplicate & Email-Matching Model
=========================================

Reconciles a marketing/prospect email list against a HubSpot contacts
export. See README.md in this folder for the full behavior spec.

Usage:
    pip install pandas openpyxl
    python hubspot_match_model.py <hubspot_export.xlsx> <target_list.xlsx> [output.xlsx]

If your target list uses different section headers than the defaults below,
edit CATEGORY_HEADERS / REGION_HEADERS / ROLE_TERMS before running.
"""

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Config — edit these if your target list uses different conventions
# ---------------------------------------------------------------------------

CATEGORY_HEADERS = {"category", "categories", "type", "segment"}
REGION_HEADERS = {"region", "territory", "area", "state", "district"}
COMPANY_HEADERS = {"company", "organization", "org", "account", "account name"}
EMAIL_HEADERS = {"email", "emails", "e-mail", "e-mails", "contact email"}

ROLE_TERMS = {
    "admin", "administrator", "info", "office", "support", "help", "sales",
    "billing", "accounts", "accounting", "hr", "careers", "jobs", "noreply",
    "no-reply", "donotreply", "webmaster", "postmaster", "parking",
    "permits", "permit", "front desk", "frontdesk", "reception",
}

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
# Loosened pattern to catch a comma-for-dot typo in the domain, e.g. name@gmail,com
EMAIL_RE_LOOSE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9,.\-]+[A-Za-z]{2,}")
# Detects two whole emails fused with no delimiter between them, e.g.
# "first@fused.comsecond@fused.com" -> captures each half so a space can be
# inserted between them. Anchored on a common-TLD ending for the FIRST email
# specifically so this never fires inside an ordinary local-part like
# "jane.doe@acme.com" (there is no second "chars@domain.tld" immediately
# following it).
_COMMON_TLDS = r"com|net|org|edu|gov|mil|io|co|biz|info|us|ca|uk"
FUSED_PAIR_RE = re.compile(
    r"([A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.(?:" + _COMMON_TLDS + r"))"
    r"([A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,})"
)
NAME_PATTERN = re.compile(r"^([A-Za-z]+)\.([A-Za-z]+)@")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize_header(col) -> str:
    return str(col).strip().lower()


def fix_typo(raw: str) -> str:
    """Fix a comma-for-dot typo in the domain part of an email-like string."""
    raw = raw.strip()
    if "@" not in raw:
        return raw
    local, _, domain = raw.partition("@")
    domain = domain.replace(",", ".")
    return f"{local}@{domain}"


def split_fused_emails(cell_text: str):
    """Extract every email address from a cell that may contain one, several
    (comma/semicolon/newline separated), or two fused together with no
    delimiter (e.g. 'a@b.comc@d.com'). Genuine duplicates are kept — dedup
    happens later via the duplicate count, not here."""
    if not isinstance(cell_text, str) or "@" not in cell_text:
        return []

    # Insert a space between any two emails fused with no delimiter so both
    # halves become independently matchable. Runs twice in case three or more
    # addresses are chained together.
    text = cell_text
    for _ in range(3):
        new_text = FUSED_PAIR_RE.sub(r"\1 \2", text)
        if new_text == text:
            break
        text = new_text

    found = []
    strict_matches = EMAIL_RE.findall(text)
    found.extend(strict_matches)

    # Blank out what strict matching already found, then look for typo'd
    # (comma-instead-of-dot domain) addresses in what's left over.
    remainder = text
    for m in strict_matches:
        remainder = remainder.replace(m, " ", 1)
    for m in EMAIL_RE_LOOSE.findall(remainder):
        found.append(fix_typo(m))

    return found


def is_role_mailbox(email: str) -> bool:
    local = email.split("@")[0].lower()
    return any(term in local for term in ROLE_TERMS)


def parse_clean_name(email: str):
    m = NAME_PATTERN.match(email)
    if not m:
        return None, None
    first, last = m.group(1), m.group(2)
    return first.capitalize(), last.capitalize()


def autosize_and_bold_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 45)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# HubSpot export loading
# ---------------------------------------------------------------------------

def load_hubspot_export(path: str) -> pd.DataFrame:
    df = pd.read_excel(path)
    df.columns = [str(c).strip() for c in df.columns]
    email_col = None
    for c in df.columns:
        if normalize_header(c) in EMAIL_HEADERS or "email" in normalize_header(c):
            email_col = c
            break
    if email_col is None:
        raise ValueError(
            "Could not find an email column in the HubSpot export. "
            "Expected a column named 'Email' or similar."
        )
    df["_match_email"] = df[email_col].astype(str).str.strip().str.lower()
    df = df[df["_match_email"].str.contains("@", na=False)]
    return df


def hubspot_known_domains(hubspot_df: pd.DataFrame) -> set:
    domains = hubspot_df["_match_email"].str.split("@").str[-1]
    return set(domains.dropna().unique())


# ---------------------------------------------------------------------------
# Target list loading — handles both the clean two-column shape and the
# grouped Category > Region > Org shape
# ---------------------------------------------------------------------------

def load_target_list(path: str):
    """Returns a list of dicts: {email, company, category, region}."""
    df = pd.read_excel(path)
    df.columns = [str(c).strip() for c in df.columns]
    norm_cols = {normalize_header(c): c for c in df.columns}

    category_col = next((norm_cols[h] for h in norm_cols if h in CATEGORY_HEADERS), None)
    region_col = next((norm_cols[h] for h in norm_cols if h in REGION_HEADERS), None)
    company_col = next((norm_cols[h] for h in norm_cols if h in COMPANY_HEADERS), None)
    email_col = next(
        (norm_cols[h] for h in norm_cols if h in EMAIL_HEADERS or "email" in h), None
    )

    records = []

    if email_col is not None and company_col is not None and category_col is None:
        # Clean two-column shape: one row per company, emails possibly packed in one cell
        for _, row in df.iterrows():
            company = row.get(company_col)
            raw_emails = row.get(email_col)
            for email in split_fused_emails(str(raw_emails)):
                records.append({
                    "email": email,
                    "company": company if pd.notna(company) else None,
                    "category": None,
                    "region": None,
                })
        return records

    # Grouped shape: forward-fill hierarchy columns, scan every other column for emails
    last_category = last_region = last_company = None
    other_cols = [c for c in df.columns if c not in {category_col, region_col, company_col}]

    for _, row in df.iterrows():
        if category_col and pd.notna(row.get(category_col)):
            last_category = row.get(category_col)
        if region_col and pd.notna(row.get(region_col)):
            last_region = row.get(region_col)
        if company_col and pd.notna(row.get(company_col)):
            last_company = row.get(company_col)

        row_company = last_company
        for col in other_cols:
            val = row.get(col)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            for email in split_fused_emails(str(val)):
                records.append({
                    "email": email,
                    "company": row_company,
                    "category": last_category,
                    "region": last_region,
                })

    return records


# ---------------------------------------------------------------------------
# Matching + workbook build
# ---------------------------------------------------------------------------

def build_workbook(hubspot_path: str, target_path: str, output_path: str):
    hubspot_df = load_hubspot_export(hubspot_path)
    known_domains = hubspot_known_domains(hubspot_df)
    target_records = load_target_list(target_path)

    if not target_records:
        raise ValueError("No email addresses were found in the target list.")

    in_hubspot_rows = []
    net_new_rows = []

    dup_counts = {}
    for rec in target_records:
        key = rec["email"].lower()
        dup_counts[key] = dup_counts.get(key, 0) + 1

    for rec in target_records:
        email = rec["email"]
        key = email.lower()
        role_flag = is_role_mailbox(email)
        match = hubspot_df[hubspot_df["_match_email"] == key]

        if not match.empty:
            hs_row = match.iloc[0].to_dict()
            hs_row.pop("_match_email", None)
            in_hubspot_rows.append({
                "Email": email,
                "Company (from list)": rec["company"],
                "Category": rec["category"],
                "Region": rec["region"],
                "Role/Shared Mailbox": role_flag,
                "Duplicate Count in List": dup_counts[key],
                **hs_row,
            })
        else:
            domain = key.split("@")[-1]
            net_new_rows.append({
                "Email": email,
                "Company (from list)": rec["company"],
                "Category": rec["category"],
                "Region": rec["region"],
                "Role/Shared Mailbox": role_flag,
                "Account Status": "Known account (new contact)" if domain in known_domains else "New account",
            })

    clean_names_rows = []
    cleanup_rows = []
    for rec in net_new_rows:
        first, last = parse_clean_name(rec["Email"])
        if first and last and not rec["Role/Shared Mailbox"]:
            clean_names_rows.append({
                "First": first,
                "Last": last,
                "Email": rec["Email"],
                "Company": rec["Company (from list)"],
            })
        else:
            if rec["Role/Shared Mailbox"]:
                confidence = "Low (role/shared mailbox)"
                hint = "Do not mail-merge a personal name; consider a generic greeting or exclude."
            elif "@" in rec["Email"] and "." not in rec["Email"].split("@")[0]:
                confidence = "Low (no first.last pattern)"
                hint = "Look up the contact name manually before importing."
            else:
                confidence = "Medium (unusual local-part format)"
                hint = "Review local-part formatting; may still be a valid first.last variant."
            cleanup_rows.append({
                "Email": rec["Email"],
                "Company": rec["Company (from list)"],
                "Confidence": confidence,
                "Hint": hint,
            })

    summary_rows = [
        {"Metric": "Total emails in list", "Count": len(target_records)},
        {"Metric": "Already in HubSpot", "Count": len(in_hubspot_rows)},
        {"Metric": "Net-new", "Count": len(net_new_rows)},
        {"Metric": "Clean names (ready to import)", "Count": len(clean_names_rows)},
        {"Metric": "Needs manual cleanup", "Count": len(cleanup_rows)},
    ]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
        pd.DataFrame(in_hubspot_rows).to_excel(writer, sheet_name="In HubSpot", index=False)
        pd.DataFrame(net_new_rows).to_excel(writer, sheet_name="Net-new", index=False)
        pd.DataFrame(clean_names_rows).to_excel(writer, sheet_name="Clean Names", index=False)
        pd.DataFrame(cleanup_rows).to_excel(writer, sheet_name="Cleanup", index=False)

        for sheet_name in ["Summary", "In HubSpot", "Net-new", "Clean Names", "Cleanup"]:
            ws = writer.sheets[sheet_name]
            if ws.max_row >= 1:
                autosize_and_bold_header(ws)

    return output_path, summary_rows


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    hubspot_path = sys.argv[1]
    target_path = sys.argv[2]
    output_path = sys.argv[3] if len(sys.argv) > 3 else "hubspot_match_output.xlsx"

    output_path, summary_rows = build_workbook(hubspot_path, target_path, output_path)

    print(f"Wrote {output_path}")
    for row in summary_rows:
        print(f"  {row['Metric']}: {row['Count']}")


if __name__ == "__main__":
    main()
